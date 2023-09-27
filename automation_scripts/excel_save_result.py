import openpyxl

def grabar_datos_inclusion_exclusion(ubicacion_operaciones, ubicacion_operaciones_mod, indice_grabar, nombre, documento, nacimiento, inicio, vencimiento, operacion, suma, costo, monto_vigente, monto_inclusion, monto_exclusion):
  """Robot que graba los montos de inclusión y exclusión en la misma planilla en una nueva hoja."""

  # Abrimos la planilla de operaciones.
  wb = openpyxl.load_workbook(ubicacion_operaciones)

  # Creamos una nueva hoja para grabar los datos.
  ws = wb.create_sheet("Datos de inclusión y exclusión")

  # Escribimos los encabezados de las columnas.
  ws.cell(row=1, column=1).value = "Nombre"
  ws.cell(row=1, column=2).value = "Nro Documento"
  ws.cell(row=1, column=3).value = "Fecha Nacimiento"
  ws.cell(row=1, column=4).value = "Fecha Inicio Operacion"
  ws.cell(row=1, column=5).value = "Fecha Vencimiento Operacion"
  ws.cell(row=1, column=6).value = "Nro Operacion"
  ws.cell(row=1, column=7).value = "Capital Recibido"
  ws.cell(row=1, column=8).value = "Costo Seguro"
  ws.cell(row=1, column=9).value = "Monto Vigente"
  ws.cell(row=1, column=10).value = "Inclusión"
  ws.cell(row=1, column=11).value = "Exclusion"

  # Escribimos los datos en la fila correspondiente.
  ws.cell(row=indice_grabar + 1, column=1).value = nombre
  ws.cell(row=indice_grabar + 1, column=2).value = documento
  ws.cell(row=indice_grabar + 1, column=3).value = nacimiento
  ws.cell(row=indice_grabar + 1, column=4).value = inicio
  ws.cell(row=indice_grabar + 1, column=5).value = vencimiento
  ws.cell(row=indice_grabar + 1, column=6).value = operacion
  ws.cell(row=indice_grabar + 1, column=7).value = suma
  ws.cell(row=indice_grabar + 1, column=8).value = costo
  ws.cell(row=indice_grabar + 1, column=9).value = monto_vigente
  ws.cell(row=indice_grabar + 1, column=10).value = monto_inclusion
  ws.cell(row=indice_grabar + 1, column=11).value = monto_exclusion

  # Guardamos la planilla de operaciones modificada.
  wb.save(ubicacion_operaciones_mod)

  # Cerramos la planilla de operaciones.
  wb.close()

# Obtenemos los datos de entrada.
ubicacion_operaciones = "C:/Users/francisco/Desktop/Excel/PlanillaDeOperaciones.xlsx"
ubicacion_operaciones_mod = "C:/Users/francisco/Desktop/Excel/PlanillaDeOperacionesModificado.xlsx"
indice_grabar = 1
nombre = "Juan Pérez"
documento = "123456789"
nacimiento = "1980-01-01"
inicio = "2023-01-01"
vencimiento = "2023-12-31"
operacion = "123456"
suma = 10000
costo = 2000
monto_vigente = 8000
monto_inclusion = 3000
monto_exclusion = 1000

# Grabamos los datos en la planilla de operaciones.
grabar_datos_inclusion_exclusion(ubicacion_operaciones, ubicacion_operaciones_mod, indice_grabar, nombre, documento, nacimiento, inicio, vencimiento, operacion, suma, costo, monto_vigente, monto_inclusion, monto_exclusion)
