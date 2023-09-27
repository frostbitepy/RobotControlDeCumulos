# DEPRECATED


import openpyxl

# Define the file path to the Excel sheet
ubicacionOperaciones = "C:/Users/francisco/Desktop/Excel/PlanillaDeOperaciones.xlsx"

def read_operations_gs(ubicacionOperaciones):
    # Initialize Excel workbook and select the active worksheet
    wb = openpyxl.load_workbook(ubicacionOperaciones)
    sheet = wb["gs"]

    # Command 2: Get the number of rows in the Excel sheet
    cantidadOperaciones = sheet.max_row - 1  # Subtract 1 for header row

    # Initialize lists to store data
    nombres = []
    nro_documento = []
    fecha_nacimiento = []
    fecha_inicio = []
    fecha_vencimiento = []
    nro_operacion = []
    suma_asegurada = []
    costo_seguro = []

    # Loop through rows and extract data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        nombres.append(row[0])
        nro_documento.append(row[1])
        fecha_nacimiento.append(row[2])
        fecha_inicio.append(row[11])
        fecha_vencimiento.append(row[12])
        nro_operacion.append(row[15])
        suma_asegurada.append(row[7])
        costo_seguro.append(row[20])

    # Close the Excel workbook
    wb.close()
    
    return {
        "cantidadOperaciones": cantidadOperaciones,
        "nombres": nombres,
        "nro_documento": nro_documento,
        "fecha_nacimiento": fecha_nacimiento,
        "fecha_inicio": fecha_inicio,
        "fecha_vencimiento": fecha_vencimiento,
        "nro_operacion": nro_operacion,
        "suma_asegurada": suma_asegurada,
        "costo_seguro": costo_seguro
    }
