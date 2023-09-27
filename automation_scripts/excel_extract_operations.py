import openpyxl


def read_operations_gs(ubicacionOperaciones):
    # Initialize Excel workbook and select the active worksheet
    wb = openpyxl.load_workbook(ubicacionOperaciones)
    sheet = wb["gs"]

    # Command 2: Get the number of rows in the Excel sheet
    cantidadOperaciones = sheet.max_row - 1  # Subtract 1 for header row

    # Initialize lists to store data
    nombres = []
    nro_documento = []
    fecha_nacimiento = []
    fecha_inicio = []
    fecha_vencimiento = []
    nro_operacion = []
    suma_asegurada = []
    costo_seguro = []

    # Loop through rows and extract data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        nombres.append(row[0])
        nro_documento.append(row[1])
        fecha_nacimiento.append(row[2])
        fecha_inicio.append(row[11])
        fecha_vencimiento.append(row[12])
        nro_operacion.append(row[15])
        suma_asegurada.append(row[7])
        costo_seguro.append(row[20])

    # Close the Excel workbook
    wb.close()
    
    return {
        "cantidadOperaciones": cantidadOperaciones,
        "nombres": nombres,
        "nro_documento": nro_documento,
        "fecha_nacimiento": fecha_nacimiento,
        "fecha_inicio": fecha_inicio,
        "fecha_vencimiento": fecha_vencimiento,
        "nro_operacion": nro_operacion,
        "suma_asegurada": suma_asegurada,
        "costo_seguro": costo_seguro
    }

def extract_operations(nombre_archivo):
    # Crea una instancia de la clase Workbook de openpyxl
    libro_excel = openpyxl.load_workbook(nombre_archivo, data_only=True)

    # Abre la primera hoja del libro
    hoja = libro_excel.active

    # Crea un diccionario donde las claves son los encabezados de columna y los valores son listas
    columnas = {hoja.cell(row=1, column=i).value: [] for i in range(1, hoja.max_column + 1)}

    # Itera a través de las filas, comenzando desde la segunda fila (omitiendo el título)
    for fila in hoja.iter_rows(values_only=True, min_row=2):
        for i, valor in enumerate(fila):
            encabezado = hoja.cell(row=1, column=i + 1).value
            columnas[encabezado].append(valor)

    # Devuelve el diccionario de columnas
    return columnas

def extract_operations_from_workbook(workbook):
    # Assume you have the workbook already open
    # Access the active sheet
    hoja = workbook.active

    # Create a dictionary where keys are column headers and values are lists
    columnas = {hoja.cell(row=1, column=i).value: [] for i in range(1, hoja.max_column + 1)}

    # Iterate through rows, starting from the second row (skipping the header)
    for fila in hoja.iter_rows(values_only=True, min_row=2):
        for i, valor in enumerate(fila):
            encabezado = hoja.cell(row=1, column=i + 1).value
            columnas[encabezado].append(valor)

    # Return the dictionary of columns
    return columnas


if __name__ == "__main__":
    extract_operations("C:/Users/francisco/Desktop/Excel/listado.xlsx")
