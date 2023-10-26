import openpyxl
import pandas as pd
from automation_scripts.script_logger import logger



def read_operations_gs(ubicacionOperaciones):
    # Initialize Excel workbook and select the active worksheet
    wb = openpyxl.load_workbook(ubicacionOperaciones)
    sheet = wb["gs"]

    # Command 2: Get the number of rows in the Excel sheet
    cantidadOperaciones = sheet.max_row - 1  # Subtract 1 for header row

    # Initialize lists to store data
    nombres = []
    nro_documento = []
    fecha_nacimiento = []
    fecha_inicio = []
    fecha_vencimiento = []
    nro_operacion = []
    suma_asegurada = []
    costo_seguro = []

    # Loop through rows and extract data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        nombres.append(row[0])
        nro_documento.append(row[1])
        fecha_nacimiento.append(row[2])
        fecha_inicio.append(row[11])
        fecha_vencimiento.append(row[12])
        nro_operacion.append(row[15])
        suma_asegurada.append(row[7])
        costo_seguro.append(row[20])

    # Close the Excel workbook
    wb.close()
    
    return {
        "cantidadOperaciones": cantidadOperaciones,
        "nombres": nombres,
        "nro_documento": nro_documento,
        "fecha_nacimiento": fecha_nacimiento,
        "fecha_inicio": fecha_inicio,
        "fecha_vencimiento": fecha_vencimiento,
        "nro_operacion": nro_operacion,
        "suma_asegurada": suma_asegurada,
        "costo_seguro": costo_seguro
    }

def extract_operations(nombre_archivo):
    # Crea una instancia de la clase Workbook de openpyxl
    libro_excel = openpyxl.load_workbook(nombre_archivo, data_only=True)

    # Abre la primera hoja del libro
    hoja = libro_excel.active

    # Crea un diccionario donde las claves son los encabezados de columna y los valores son listas
    columnas = {hoja.cell(row=1, column=i).value: [] for i in range(1, hoja.max_column + 1)}

    # Itera a través de las filas, comenzando desde la segunda fila (omitiendo el título)
    for fila in hoja.iter_rows(values_only=True, min_row=2):
        for i, valor in enumerate(fila):
            encabezado = hoja.cell(row=1, column=i + 1).value
            columnas[encabezado].append(valor)

    # Devuelve el diccionario de columnas
    logger.info("Operaciones extraidas")
    return columnas

def extract_operations_from_workbook(workbook):
    # Assume you have the workbook already open
    # Access the active sheet
    hoja = workbook.active

    # Create a dictionary where keys are column headers and values are lists
    columnas = {hoja.cell(row=1, column=i).value: [] for i in range(1, hoja.max_column + 1)}

    # Iterate through rows, starting from the second row (skipping the header)
    for fila in hoja.iter_rows(values_only=True, min_row=2):
        for i, valor in enumerate(fila):
            encabezado = hoja.cell(row=1, column=i + 1).value
            columnas[encabezado].append(valor)

    # Return the dictionary of columns
    logger.info("Operaciones extraidas")
    return columnas


def generate_report(cliente, documento, fecha_nacimiento, operacion, fecha_inicio, fecha_vencimiento, suma_asegurada):
    # Crear un DataFrame de Pandas con los datos
    data = {
        'Cliente': [cliente],
        'Documento': [documento],
        'Fecha de Nacimiento': [fecha_nacimiento],
        'Operación': [operacion],
        'Fecha de Inicio': [fecha_inicio],
        'Fecha de Vencimiento': [fecha_vencimiento],
        'Suma Asegurada': [suma_asegurada]
    }

    df = pd.DataFrame(data)

    # Guardar el DataFrame en un archivo Excel
    writer = pd.ExcelWriter('datos_cliente.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Datos', index=False)

    # Configurar el formato de las celdas si es necesario
    # workbook = writer.book
    # worksheet = writer.sheets['Datos']
    # format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # worksheet.set_column('C:C', None, format)

    # Cerrar el archivo Excel
    logger.info("Archivo de datos del cliente generado")
    writer.save()


if __name__ == "__main__":
    extract_operations("listado.xlsx")
